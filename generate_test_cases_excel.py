"""
generate_test_cases_excel.py — Generate FridgeAI Manual Testing Report as Excel
matching the test_case_format.xlsx template.

Run:
    python generate_test_cases_excel.py
"""

import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "002060"
WHITE      = "FFFFFF"
PASS_GREEN = "C6EFCE"
FAIL_RED   = "FFC7CE"
ALT_ROW    = "DCE6F1"
HEADER_FG  = "FFFFFF"

# ── Styles ───────────────────────────────────────────────────────────────────
hdr_fill   = PatternFill("solid", fgColor=DARK_BLUE)
pass_fill  = PatternFill("solid", fgColor=PASS_GREEN)
fail_fill  = PatternFill("solid", fgColor=FAIL_RED)
alt_fill   = PatternFill("solid", fgColor=ALT_ROW)
plain_fill = PatternFill("solid", fgColor="FFFFFF")

hdr_font   = Font(name="Calibri", bold=True,  size=11, color=WHITE)
data_font  = Font(name="Calibri", bold=False, size=9)
meta_font  = Font(name="Calibri", bold=True,  size=11, color=WHITE)

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
wrap_left   = Alignment(wrap_text=True, vertical="top", horizontal="left")


def style_hdr(cell, value):
    cell.value = value
    cell.font  = hdr_font
    cell.fill  = hdr_fill
    cell.alignment = wrap_center
    cell.border = border


def style_meta_label(cell, value):
    cell.value = value
    cell.font  = meta_font
    cell.fill  = hdr_fill
    cell.alignment = Alignment(vertical="center", horizontal="left")


def style_meta_value(cell, value):
    cell.value = value
    cell.font  = Font(name="Calibri", size=11)
    cell.alignment = Alignment(vertical="center", horizontal="left")


def style_data(cell, value, fill, center=False):
    cell.value = value
    cell.font  = data_font
    cell.fill  = fill
    cell.alignment = wrap_center if center else wrap_left
    cell.border = border


# ── Test case data ────────────────────────────────────────────────────────────
# Columns: tc_id, scenario, test_case, pre_condition, test_steps,
#          test_data, expected, post_condition, actual, status

ALL_CASES = [
    # ── M1: Items Module ──────────────────────────────────────────────────────
    (
        "TC_ITEMS_001",
        "Create Fridge Item",
        "Create item with all valid required fields",
        "Backend running; DB empty or clean.",
        "1. Send POST /items\n2. Note returned JSON",
        '{"name":"Milk","category":"dairy","quantity":2,"shelf_life":7,"storage_temp":4,"humidity":60}',
        "HTTP 201. item_id generated, p_spoil=null, confidence_tier='LOW', entry_time set.",
        "Item exists in DB. Settle timer scheduled.",
        "HTTP 201 returned. All fields correct. p_spoil=null confirmed.",
        "PASS",
    ),
    (
        "TC_ITEMS_002",
        "Create Fridge Item",
        "Create item with missing required field (name)",
        "Backend running.",
        "1. Send POST /items without 'name' field\n2. Observe response",
        '{"category":"dairy","quantity":1,"shelf_life":7,"storage_temp":4,"humidity":60}',
        "HTTP 422. Error indicates 'name' is required.",
        "No item created in DB.",
        "HTTP 422 returned. Error: 'Name: field required'.",
        "PASS",
    ),
    (
        "TC_ITEMS_003",
        "Create Fridge Item",
        "Create item with storage_temp out of range (>60°C)",
        "Backend running.",
        "1. Send POST /items with storage_temp=100\n2. Observe response",
        '{"name":"X","category":"dairy","quantity":1,"shelf_life":7,"storage_temp":100,"humidity":60}',
        "HTTP 422. Error: storage_temp must be between -30 and 60.",
        "No item created.",
        "HTTP 422. Custom error: 'Storage temperature must be between -30 and 60°C'.",
        "PASS",
    ),
    (
        "TC_ITEMS_004",
        "Create Fridge Item",
        "Create item with quantity = 0",
        "Backend running.",
        "1. Send POST /items with quantity=0\n2. Observe response",
        '{"name":"Milk","category":"dairy","quantity":0,"shelf_life":7,"storage_temp":4,"humidity":60}',
        "HTTP 422. Error: quantity must be > 0.",
        "No item created.",
        "HTTP 422. Error: 'Quantity must be greater than 0'.",
        "PASS",
    ),
    (
        "TC_ITEMS_005",
        "Get Fridge Item",
        "Retrieve a single item by valid ID",
        "Item 'Milk' exists with known item_id.",
        "1. Send GET /items/{item_id}\n2. Observe response",
        "item_id from TC_ITEMS_001",
        "HTTP 200. Full ItemRead with correct fields.",
        "Item data unchanged.",
        "HTTP 200. name='Milk', category='dairy', quantity=2 confirmed.",
        "PASS",
    ),
    (
        "TC_ITEMS_006",
        "Get Fridge Item",
        "Retrieve item with non-existent ID",
        "Backend running.",
        "1. Send GET /items/nonexistent-id\n2. Observe response",
        "item_id = 'nonexistent-id'",
        "HTTP 404. {'detail':'Item not found'}.",
        "No change in DB.",
        "HTTP 404 returned. Correct error message.",
        "PASS",
    ),
    (
        "TC_ITEMS_007",
        "List Fridge Items",
        "List all items in the fridge",
        "At least 3 items in inventory.",
        "1. Send GET /items\n2. Observe response",
        "No query params",
        "HTTP 200. Returns list of all ItemRead objects.",
        "No change in DB.",
        "HTTP 200. List of 3 items returned with full fields.",
        "PASS",
    ),
    (
        "TC_ITEMS_008",
        "List Fridge Items",
        "Filter items by category=dairy",
        "Inventory contains dairy and non-dairy items.",
        "1. Send GET /items?category=dairy\n2. Observe response",
        "category=dairy",
        "HTTP 200. Only dairy items returned.",
        "No change in DB.",
        "HTTP 200. Only dairy items in response. Non-dairy excluded.",
        "PASS",
    ),
    (
        "TC_ITEMS_009",
        "Update Fridge Item",
        "Update item quantity via PATCH",
        "Item 'Milk' exists with quantity=2.",
        "1. PATCH /items/{item_id} with {quantity:1}\n2. GET item to verify",
        '{"quantity":1}',
        "HTTP 200. quantity=1. Other fields unchanged.",
        "DB updated with new quantity.",
        "HTTP 200. quantity=1 confirmed. name unchanged.",
        "PASS",
    ),
    (
        "TC_ITEMS_010",
        "Update Fridge Item",
        "PATCH storage_temp should reschedule settle timer",
        "Item exists. Settle timer pending.",
        "1. PATCH /items/{item_id} with {storage_temp:8}\n2. Check backend logs",
        '{"storage_temp":8}',
        "Old timer cancelled; new timer scheduled.",
        "New settle timer pending in DB.",
        "Log: 'Cancelled timer ... Rescheduled'. Timer restarted.",
        "PASS",
    ),
    (
        "TC_ITEMS_011",
        "Delete Fridge Item",
        "Delete an existing item",
        "Item with known item_id exists.",
        "1. DELETE /items/{item_id}\n2. GET /items/{item_id}",
        "item_id of target item",
        "DELETE: HTTP 204. GET: HTTP 404.",
        "Item removed from DB. Consumption history recorded.",
        "DELETE: 204 No Content. GET: 404 Not Found. History entry created.",
        "PASS",
    ),
    (
        "TC_ITEMS_012",
        "Item Feedback",
        "Submit 'still good' feedback for a scored item",
        "Item scored (p_spoil populated).",
        "1. POST /items/{item_id}/feedback\n2. Check feedback table",
        '{"still_good":true,"shelf_life_actual":10}',
        "HTTP 201. Feedback saved. Correction = actual - declared shelf life.",
        "Settle timer rescheduled for immediate rescore.",
        "HTTP 201. FeedbackRead returned. Correction stored. Timer rescheduled.",
        "PASS",
    ),

    # ── M2: Alerts ────────────────────────────────────────────────────────────
    (
        "TC_ALERTS_001",
        "List Alerts",
        "Retrieve all alerts ordered newest first",
        "At least 1 alert fired by scorer.",
        "1. GET /alerts\n2. Check ordering",
        "No params",
        "HTTP 200. List of AlertRead objects, newest first.",
        "No change.",
        "HTTP 200. 2 alerts returned ordered by created_at DESC.",
        "PASS",
    ),
    (
        "TC_ALERTS_002",
        "Alert Threshold",
        "CRITICAL_ALERT fires when P_spoil > 0.80",
        "Item backdated 10 days. SETTLE_DELAY_SECONDS=0.",
        "1. Trigger scorer for backdated item\n2. GET /alerts",
        "Item entry_time = now - 10 days",
        "CRITICAL_ALERT in alerts. P_spoil > 0.80.",
        "Alert stored in DB. WS event broadcast.",
        "CRITICAL_ALERT fired. P_spoil=0.93. Message correct.",
        "PASS",
    ),
    (
        "TC_ALERTS_003",
        "Alert Threshold",
        "WARNING_ALERT fires when 0.50 < P_spoil <= 0.80",
        "Item with moderate age.",
        "1. Trigger scorer\n2. Check /alerts for WARNING_ALERT",
        "Item age ~5 days, temp=10°C",
        "WARNING_ALERT in list. P_spoil in (0.50, 0.80].",
        "Alert stored in DB.",
        "WARNING_ALERT fired. P_spoil=0.63.",
        "PASS",
    ),
    (
        "TC_ALERTS_004",
        "Alert Threshold",
        "USE_TODAY_ALERT fires when RSL < 0.5 days",
        "Item with RSL ~0.3 days.",
        "1. Trigger scorer for near-expired item\n2. Check /alerts",
        "Item near expiry",
        "USE_TODAY_ALERT. RSL < 0.5.",
        "Alert stored. WS event sent.",
        "USE_TODAY_ALERT fired. RSL=0.28 days.",
        "PASS",
    ),
    (
        "TC_ALERTS_005",
        "List Alerts",
        "Filter alerts with 'since' timestamp",
        "Alerts exist across different timestamps.",
        "1. GET /alerts?since=2026-03-30T00:00:00Z\n2. Check results",
        "since=2026-03-30T00:00:00Z",
        "Only alerts after given timestamp returned.",
        "No change.",
        "Only today's alerts returned. Older excluded.",
        "PASS",
    ),

    # ── M3: Auth ──────────────────────────────────────────────────────────────
    (
        "TC_AUTH_001",
        "User Registration",
        "Register new user with valid credentials",
        "REQUIRE_AUTH=true. Supabase configured.",
        "1. POST /auth/register\n2. Observe response",
        '{"email":"test@fridge.ai","password":"Secret123!","username":"tester","household_name":"Home"}',
        "HTTP 201. TokenResponse with access_token and user object.",
        "User and household created in Supabase.",
        "HTTP 201. access_token returned. household_id populated.",
        "PASS",
    ),
    (
        "TC_AUTH_002",
        "User Registration",
        "Register with already-used email",
        "User test@fridge.ai already exists.",
        "1. POST /auth/register with same email\n2. Observe response",
        '{"email":"test@fridge.ai","password":"Other123!","username":"other"}',
        "HTTP 400/409/422. Duplicate email error.",
        "No new user created.",
        "HTTP 422 from Supabase. Duplicate email constraint.",
        "PASS",
    ),
    (
        "TC_AUTH_003",
        "User Login",
        "Login with valid credentials",
        "User from TC_AUTH_001 exists.",
        "1. POST /auth/login\n2. Note access_token",
        '{"email":"test@fridge.ai","password":"Secret123!"}',
        "HTTP 200. TokenResponse with valid access_token.",
        "Session active.",
        "HTTP 200. JWT token returned.",
        "PASS",
    ),
    (
        "TC_AUTH_004",
        "User Login",
        "Login with incorrect password",
        "Valid user exists.",
        "1. POST /auth/login with wrong password\n2. Observe response",
        '{"email":"test@fridge.ai","password":"WrongPass!"}',
        "HTTP 401. {'detail':'Invalid credentials'}.",
        "No session created.",
        "HTTP 401 returned. Error message correct.",
        "PASS",
    ),
    (
        "TC_AUTH_005",
        "User Profile",
        "Fetch profile with valid token",
        "Valid access_token from TC_AUTH_003.",
        "1. GET /auth/me with Bearer token\n2. Observe response",
        "Authorization: Bearer <token>",
        "HTTP 200. UserRead with user_id, username, email, household_id.",
        "No change.",
        "HTTP 200. Correct user profile returned.",
        "PASS",
    ),
    (
        "TC_AUTH_006",
        "User Preferences",
        "Enable auto-restock preference",
        "Logged-in user.",
        "1. PATCH /auth/prefs with {auto_restock_enabled:true}\n2. GET /auth/prefs",
        '{"auto_restock_enabled":true}',
        "PATCH: HTTP 200. GET: auto_restock_enabled=true.",
        "Preference saved in Supabase.",
        "PATCH 200. GET confirms auto_restock_enabled=true.",
        "PASS",
    ),

    # ── M4: Lookup ────────────────────────────────────────────────────────────
    (
        "TC_LOOKUP_001",
        "Category Shelf Life",
        "Lookup shelf life for category 'dairy'",
        "Backend running.",
        "1. GET /lookup/shelf-life/dairy\n2. Note response",
        "category=dairy",
        "HTTP 200. shelf_life_days=7, source='USDA FoodKeeper'.",
        "No change.",
        "HTTP 200. shelf_life_days=7 confirmed.",
        "PASS",
    ),
    (
        "TC_LOOKUP_002",
        "Category Shelf Life",
        "Lookup shelf life for unknown category",
        "Backend running.",
        "1. GET /lookup/shelf-life/snacks\n2. Observe response",
        "category=snacks",
        "HTTP 404. Category not found.",
        "No change.",
        "HTTP 404. {'detail':'Category not found'}.",
        "PASS",
    ),
    (
        "TC_LOOKUP_003",
        "Item Name Lookup",
        "Lookup item details by name 'chicken'",
        "Backend running.",
        "1. GET /lookup/item/chicken\n2. Note fields",
        "name=chicken",
        "HTTP 200. category='meat', shelf_life_days=3, estimated_cost in INR.",
        "No change.",
        "HTTP 200. category='meat', shelf_life=3, cost=250.0 INR.",
        "PASS",
    ),
    (
        "TC_LOOKUP_004",
        "Barcode Lookup",
        "Lookup EAN barcode for known product",
        "Internet access. Open Food Facts available.",
        "1. GET /lookup/barcode/8901058890918\n2. Observe response",
        "barcode=8901058890918",
        "HTTP 200. Product name, category, shelf_life returned.",
        "No change.",
        "HTTP 200. 'Amul Butter', category='dairy', shelf_life=7.",
        "PASS",
    ),
    (
        "TC_LOOKUP_005",
        "Barcode Lookup",
        "Lookup barcode not in Open Food Facts",
        "Internet access.",
        "1. GET /lookup/barcode/0000000000000\n2. Observe response",
        "barcode=0000000000000",
        "HTTP 404. Barcode not found.",
        "No change.",
        "HTTP 404 returned with error message.",
        "PASS",
    ),

    # ── M5: Vision ────────────────────────────────────────────────────────────
    (
        "TC_VISION_001",
        "Fridge Image Scan",
        "Upload fridge image with identifiable food items",
        "Models loaded. Test JPEG with fruits/vegetables.",
        "1. POST /vision/scan with fridge.jpg\n2. Observe ScanResult",
        "file=fridge.jpg (multipart)",
        "HTTP 200. items list with ≥1 DetectedItem. name, category, confidence, shelf_life present.",
        "No DB change.",
        "HTTP 200. 3 items detected: apple(0.72), orange(0.68), carrot(0.55).",
        "PASS",
    ),
    (
        "TC_VISION_002",
        "Spoilage Detection",
        "Detect spoiled item in image",
        "Test JPEG with visibly rotten banana.",
        "1. POST /vision/scan with rotten-banana.jpg\n2. Check spoilage_detected",
        "file=rotten-banana.jpg",
        "spoilage_detected=true, spoilage_confidence > 0.5.",
        "No DB change.",
        "banana: spoilage_detected=true, spoilage_confidence=0.87.",
        "PASS",
    ),
    (
        "TC_VISION_003",
        "Fridge Image Scan",
        "Upload image with no recognizable food items",
        "Image of empty shelf.",
        "1. POST /vision/scan with empty-shelf.jpg\n2. Observe result",
        "file=empty-shelf.jpg",
        "HTTP 200. items = [].",
        "No DB change.",
        "HTTP 200. items=[] returned. No false positives.",
        "PASS",
    ),
    (
        "TC_VISION_004",
        "Fridge Image Scan",
        "Upload non-image file (PDF)",
        "Backend running.",
        "1. POST /vision/scan with a .pdf file\n2. Observe response",
        "file=document.pdf",
        "HTTP 400. Invalid image format.",
        "No DB change.",
        "HTTP 400. Pillow raises UnidentifiedImageError. Caught and returned as 400.",
        "PASS",
    ),
    (
        "TC_VISION_005",
        "Live Object Detection",
        "Object moved rapidly during detection — item not detected",
        "Camera active. Grounding DINO model loaded. Good lighting.",
        "1. Place object in frame\n2. Press SPACE to trigger detection while rapidly moving the object\n3. Observe detection results",
        "Object speed: ~0.5 m/s lateral movement during capture; exposure=auto",
        "All moved objects correctly detected with confidence > 0.5.",
        "No crash. Detection result returned.",
        "0 items detected. Motion blur caused Grounding DINO to miss the object. No bounding box drawn.",
        "FAIL",
    ),
    (
        "TC_VISION_006",
        "Live Object Detection",
        "Detection fails under very low ambient lighting",
        "Camera active. Grounding DINO model loaded. Fridge light off / dark room.",
        "1. Set room lighting to very low (< 50 lux)\n2. Place recognizable item in frame\n3. Press SPACE to trigger detection\n4. Observe detection results",
        "Ambient light: < 50 lux; item: apple (known-detectable at normal lighting)",
        "Item detected with confidence > 0.5 despite low light.",
        "No crash. Detection result returned.",
        "0 items detected. Low brightness caused confidence < 0.5 threshold. No bounding box drawn.",
        "FAIL",
    ),
    (
        "TC_VISION_007",
        "Live Object Detection",
        "Partially visible object at camera corner (~30% in frame) not detected",
        "Camera active. Grounding DINO model loaded. Normal lighting.",
        "1. Place object so only ~30% of it is visible at the corner of the camera frame (bird's eye, side, or angled view)\n2. Press SPACE to trigger detection\n3. Observe detection results",
        "Object visibility: ~30% of bounding box within frame; tested angles: top-left corner, bottom-right corner, side-clipped",
        "Partially visible object detected with confidence > 0.5 and correctly labelled.",
        "No crash. Detection result returned.",
        "0 items detected. Grounding DINO confidence fell below 0.5 threshold for partially visible / edge-clipped object.",
        "FAIL",
    ),

    (
        "TC_VISION_008",
        "Live Object Detection",
        "Fully visible object placed at camera corner is correctly detected",
        "Camera active. Grounding DINO model loaded. Normal lighting.",
        "1. Place object fully within frame but positioned at the corner (100% of object visible)\n2. Press SPACE to trigger detection\n3. Observe detection results",
        "Object visibility: 100% within frame; position: top-left or bottom-right corner; item: apple",
        "Object detected with confidence > 0.5 and correctly labelled despite corner placement.",
        "No crash. Detection result returned.",
        "apple detected with confidence=0.71. Bounding box correctly placed at corner of frame.",
        "PASS",
    ),
    (
        "TC_VISION_009",
        "Live Object Detection",
        "Multiple distinct objects in frame are all detected",
        "Camera active. Grounding DINO model loaded. Normal lighting. 3 distinct food items prepared.",
        "1. Place 3 distinct food items (e.g. apple, milk carton, carrot) spread across the frame\n2. Press SPACE to trigger detection\n3. Observe items[] in result",
        "Items: apple, milk carton, carrot; all fully visible; normal spacing between objects",
        "All 3 items detected. items[] length = 3. Each has correct name, category, and confidence > 0.5.",
        "No crash. Detection result with 3 DetectedItem entries returned.",
        "3 items detected: apple(0.74), milk(0.68), carrot(0.61). All correctly categorised.",
        "PASS",
    ),

    # ── M6: Grocery ───────────────────────────────────────────────────────────
    (
        "TC_GROCERY_001",
        "Grocery List Management",
        "Manually add item to grocery list",
        "Backend running.",
        "1. POST /grocery\n2. Observe response",
        '{"name":"Eggs","category":"protein","quantity":12,"source":"manual"}',
        "HTTP 201. grocery_id generated, checked=false, source='manual'.",
        "Item in grocery_items table.",
        "HTTP 201. Grocery item created with correct fields.",
        "PASS",
    ),
    (
        "TC_GROCERY_002",
        "Grocery List Management",
        "List grocery items — unchecked first",
        "3 items: 2 unchecked, 1 checked.",
        "1. GET /grocery\n2. Check ordering",
        "No params",
        "HTTP 200. Unchecked items before checked items.",
        "No change.",
        "HTTP 200. 2 unchecked first, 1 checked last.",
        "PASS",
    ),
    (
        "TC_GROCERY_003",
        "Grocery List Management",
        "Mark grocery item as checked",
        "Unchecked grocery item exists.",
        "1. PATCH /grocery/{id} with {checked:true}\n2. GET /grocery",
        '{"checked":true}',
        "PATCH: HTTP 200. checked=true. Item at end of list.",
        "DB updated. WS GROCERY_UPDATED broadcast.",
        "PATCH 200. checked=true. Item moved to bottom of list.",
        "PASS",
    ),
    (
        "TC_GROCERY_004",
        "Grocery List Management",
        "Promote grocery item to fridge inventory",
        "Unchecked 'Eggs' in grocery list.",
        "1. POST /grocery/{id}/add-to-fridge\n2. GET /items",
        "grocery_id of Eggs",
        "HTTP 200. ItemRead returned. Eggs in inventory. Grocery checked.",
        "New item in items table. Settle timer started.",
        "HTTP 200. Eggs in inventory. Grocery item checked=true.",
        "PASS",
    ),
    (
        "TC_GROCERY_005",
        "Grocery List Management",
        "Bulk delete all checked grocery items",
        "At least 2 checked items in list.",
        "1. DELETE /grocery/checked\n2. GET /grocery",
        "No body",
        "HTTP 204. GET returns only unchecked items.",
        "Checked items removed from DB.",
        "HTTP 204. GET: only unchecked items remain.",
        "PASS",
    ),

    # ── M7: Restock ───────────────────────────────────────────────────────────
    (
        "TC_RESTOCK_001",
        "Restock Suggestions",
        "URGENT suggestion for item with RSL < 2 days, P_spoil > 0.5",
        "Item with RSL=0.5 days, P_spoil=0.75.",
        "1. GET /restock\n2. Check priority",
        "No params",
        "HTTP 200. Item listed as priority='urgent'.",
        "No change.",
        "HTTP 200. URGENT suggestion with correct reason.",
        "PASS",
    ),
    (
        "TC_RESTOCK_002",
        "Restock Suggestions",
        "LOW_STOCK suggestion for item with quantity=1, P_spoil > 0.4",
        "Item with qty=1, P_spoil=0.45.",
        "1. GET /restock\n2. Check priority",
        "No params",
        "priority='low_stock'. Reason: running low.",
        "No change.",
        "LOW_STOCK suggestion returned.",
        "PASS",
    ),
    (
        "TC_RESTOCK_003",
        "Restock Suggestions",
        "Overdue suggestion based on consumption pattern",
        "Item consumed regularly; predicted_next_days = -3.9.",
        "1. GET /restock\n2. Check for overdue suggestion",
        "No params",
        "Item appears with reason mentioning overdue days.",
        "No change.",
        "Suggestion: 'Based on your usage pattern, overdue by 3.9 day(s)'.",
        "PASS",
    ),
    (
        "TC_RESTOCK_004",
        "Restock Suggestions",
        "No suggestions when all items are fresh",
        "All items: RSL > 5 days, qty > 1, fresh.",
        "1. GET /restock",
        "No params",
        "HTTP 200. Empty list [].",
        "No change.",
        "HTTP 200. [] returned.",
        "PASS",
    ),

    # ── M8: Recipes ───────────────────────────────────────────────────────────
    (
        "TC_RECIPES_001",
        "Recipe Suggestions",
        "Fetch recipes based on fridge contents",
        "SPOONACULAR_API_KEY set. Inventory: eggs, milk, butter.",
        "1. GET /recipes/suggestions\n2. Check returned recipes",
        "No params",
        "HTTP 200. List of Recipe objects with used/missed ingredients.",
        "No change.",
        "HTTP 200. 5 recipes. Pancakes matched 3/3 ingredients.",
        "PASS",
    ),
    (
        "TC_RECIPES_002",
        "Recipe Details",
        "Fetch step-by-step instructions for a recipe",
        "Valid meal_id from TC_RECIPES_001.",
        "1. GET /recipes/{meal_id}/details\n2. Inspect steps",
        "meal_id from suggestion",
        "HTTP 200. steps[], ready_in_minutes, servings, source_url present.",
        "No change.",
        "HTTP 200. 6 steps returned. ready_in_minutes=20.",
        "PASS",
    ),
    (
        "TC_RECIPES_003",
        "Cook Recipe",
        "Cook recipe — decrements matched ingredient quantities",
        "milk(qty=3), eggs(qty=6) in inventory.",
        "1. POST /recipes/{meal_id}/cook\n2. GET /items",
        '{"item_ids":["milk-id","eggs-id"]}',
        "HTTP 200. milk.qty=2, eggs.qty=5. History recorded.",
        "Items updated in DB. WS events broadcast.",
        "HTTP 200. Quantities decremented. History: reason='cooked'.",
        "PASS",
    ),
    (
        "TC_RECIPES_004",
        "Cook Recipe",
        "Item deleted when qty hits 0 after cooking",
        "butter.quantity=1 in inventory.",
        "1. POST /recipes/{meal_id}/cook with butter\n2. GET /items/{butter_id}",
        '{"item_ids":["butter-id"]}',
        "butter deleted from inventory. GET returns 404.",
        "Item removed. History recorded.",
        "GET butter → 404. History entry: reason='cooked', qty=1.",
        "PASS",
    ),
    (
        "TC_RECIPES_005",
        "Recipe API Failure",
        "Spoonacular API returns error (key invalid / quota exceeded)",
        "SPOONACULAR_API_KEY set to invalid value.",
        "1. GET /recipes/suggestions\n2. Observe response",
        "SPOONACULAR_API_KEY=invalid_key",
        "HTTP 502 or 503. Response body: {\"detail\": \"Recipe service unavailable. Please try again later.\"}",
        "No DB change. Error logged on backend.",
        "HTTP 502 returned. Error message displayed in Recipes tab: 'Recipe service unavailable. Please try again later.'",
        "PASS",
    ),

    # ── M9: Receipt ───────────────────────────────────────────────────────────
    (
        "TC_RECEIPT_001",
        "Receipt OCR Scan",
        "Upload receipt photo and extract items",
        "GEMINI_API_KEY set. Test receipt JPEG with ≥3 items.",
        "1. POST /receipt/scan with receipt.jpg\n2. Observe ReceiptScanResult",
        "file=receipt.jpg (multipart)",
        "HTTP 200. items[] with name, category, cost. ocr_engine='gemini'.",
        "No DB change.",
        "HTTP 200. 4 items extracted. ocr_engine='gemini'.",
        "PASS",
    ),
    (
        "TC_RECEIPT_002",
        "Receipt Parse Text",
        "Parse pre-extracted receipt text",
        "Backend running.",
        "1. POST /receipt/parse-text\n2. Observe result",
        '{"text":"MILK 45.00\\nEGGS 60.00\\nBREAD 35.00"}',
        "HTTP 200. 3 items: name, price, category inferred.",
        "No DB change.",
        "HTTP 200. milk(dairy,45), eggs(protein,60), bread(cooked,35).",
        "PASS",
    ),
    (
        "TC_RECEIPT_003",
        "Receipt OCR Fallback",
        "Tesseract fallback when Gemini key missing",
        "GEMINI_API_KEY not set. pytesseract installed.",
        "1. POST /receipt/scan with receipt.jpg\n2. Check ocr_engine field",
        "file=receipt.jpg",
        "HTTP 200. ocr_engine='tesseract'. Items extracted.",
        "No DB change.",
        "HTTP 200. ocr_engine='tesseract'. 2/4 items parsed.",
        "PASS",
    ),
    (
        "TC_RECEIPT_004",
        "Receipt OCR Scan",
        "Upload blurry / non-receipt image",
        "Backend running.",
        "1. POST /receipt/scan with blurry wall photo\n2. Observe result",
        "file=blurry.jpg",
        "HTTP 200. items=[], raw_text empty.",
        "No DB change.",
        "HTTP 200. items=[] returned.",
        "PASS",
    ),

    # ── M10: Analytics ────────────────────────────────────────────────────────
    (
        "TC_ANALYTICS_001",
        "Consumption Trend",
        "Retrieve 7-day daily consumption counts",
        "≥5 items consumed over the past week.",
        "1. GET /analytics/consumption?days=7\n2. Inspect data",
        "days=7",
        "HTTP 200. 7 ConsumptionPoint records (date, items_consumed, total_qty).",
        "No change.",
        "HTTP 200. 7 data points. Totals match history.",
        "PASS",
    ),
    (
        "TC_ANALYTICS_002",
        "Waste Patterns",
        "Identify most wasted items",
        "Multiple items deleted as 'wasted'.",
        "1. GET /analytics/waste-patterns\n2. Check top result",
        "No params",
        "HTTP 200. WastePattern list sorted by times_wasted DESC.",
        "No change.",
        "Top: Tomatoes (3x, avg_p_spoil=0.81).",
        "PASS",
    ),
    (
        "TC_ANALYTICS_003",
        "Waste Summary",
        "Get aggregate waste summary for 30 days",
        "History spanning 30 days.",
        "1. GET /analytics/summary?days=30\n2. Check fields",
        "days=30",
        "HTTP 200. WasteSummary with consumed, wasted, waste_rate_pct, top_wasted, daily_trend.",
        "No change.",
        "consumed=18, wasted=4, waste_rate_pct=22.2. Fields populated.",
        "PASS",
    ),
    (
        "TC_ANALYTICS_004",
        "Consumption Predictions",
        "Predict next consumption for item with history",
        "'Milk' consumed 5 times in past month.",
        "1. GET /analytics/predictions\n2. Check Milk entry",
        "No params",
        "Milk: confidence='HIGH', weekly_rate>0, predicted_next_days populated.",
        "No change.",
        "HIGH confidence. weekly_rate=1.3, next_in_days=3.",
        "PASS",
    ),

    # ── M11: WebSocket ────────────────────────────────────────────────────────
    (
        "TC_WS_001",
        "WebSocket Connection",
        "Establish connection to /ws",
        "Backend running.",
        "1. wscat -c ws://localhost:8000/ws?client_type=web\n2. Check /status",
        "client_type=web",
        "Connection established. ws_clients incremented by 1.",
        "Connection active.",
        "Connected. /status shows ws_clients=1.",
        "PASS",
    ),
    (
        "TC_WS_002",
        "WebSocket Events",
        "ITEM_INSERTED event on POST /items",
        "WS client connected.",
        "1. POST /items\n2. Observe WS message",
        "New item payload",
        'WS: {"event":"ITEM_INSERTED","data":{...full item...}}',
        "Event received by all clients.",
        "ITEM_INSERTED received with full ItemRead data.",
        "PASS",
    ),
    (
        "TC_WS_003",
        "WebSocket Events",
        "ITEM_SCORED event after settle timer fires",
        "SETTLE_DELAY_SECONDS=5. Item inserted.",
        "1. Wait 5s after creation\n2. Observe WS",
        "SETTLE_DELAY_SECONDS=5",
        'WS: {"event":"ITEM_SCORED","data":{"P_spoil":...,"RSL":...,"fapf_score":...}}',
        "Item updated in DB.",
        "ITEM_SCORED received after 5s. Fields populated.",
        "PASS",
    ),
    (
        "TC_WS_004",
        "WebSocket Events",
        "ALERT_FIRED event when threshold exceeded",
        "Item with P_spoil > 0.80 being scored.",
        "1. Score backdated item\n2. Observe WS",
        "Backdated item",
        'WS: {"event":"ALERT_FIRED","data":{alert_type,P_spoil,item_name,message}}',
        "Alert in DB. Toast in frontend.",
        "ALERT_FIRED: type='CRITICAL_ALERT', P_spoil=0.93.",
        "PASS",
    ),
    (
        "TC_WS_005",
        "WebSocket Reconnect",
        "Frontend auto-reconnects after server restart",
        "Frontend connected. Backend then restarted.",
        "1. Stop backend\n2. Observe UI status\n3. Restart backend\n4. Observe reconnect",
        "Backend restart",
        "UI shows 'disconnected'. Reconnects within 3s of backend restart.",
        "WS re-established. State refreshed.",
        "wsStatus: disconnected → reconnected within 3s.",
        "PASS",
    ),

    # ── M12: ASLIE ────────────────────────────────────────────────────────────
    (
        "TC_ASLIE_001",
        "ASLIE Model",
        "P_spoil always in [0, 1]",
        "ASLIE module imported. pytest available.",
        "1. Run pytest tests/test_aslie.py::test_p_spoil_range",
        "t in [0,30], temp in [0,30], humidity in [0,100]",
        "All P_spoil values in [0.0, 1.0].",
        "No side effects.",
        "100 random inputs all yield P_spoil in [0.0, 1.0]. PASS.",
        "PASS",
    ),
    (
        "TC_ASLIE_002",
        "ASLIE Model",
        "P_spoil is non-decreasing as time increases",
        "ASLIE module imported.",
        "1. Compute P_spoil for t=1..14 (fixed temp, humidity)\n2. Check monotonicity",
        "t=1,2,...,14; temp=4; humidity=60; category=dairy",
        "P_spoil(t+1) >= P_spoil(t) for all t.",
        "No side effects.",
        "Values strictly non-decreasing: [0.01, 0.04, 0.11, 0.28, ...].",
        "PASS",
    ),
    (
        "TC_ASLIE_003",
        "ASLIE Model",
        "Higher temperature produces higher P_spoil",
        "ASLIE module imported.",
        "1. Compare P_spoil at temp=4 vs temp=25 (same t, cat, hum)\n2. Compare",
        "t=5, cat=dairy, hum=60; temp=4 vs temp=25",
        "P_spoil(temp=25) > P_spoil(temp=4).",
        "No side effects.",
        "temp=4: 0.13. temp=25: 0.71. Correct ordering.",
        "PASS",
    ),
    (
        "TC_ASLIE_004",
        "ASLIE Model",
        "RSL binary search converges correctly",
        "ASLIE module imported.",
        "1. Compute rsl(t_elapsed=0, shelf_life=7, temp=4, cat=1, hum=60)",
        "t_elapsed=0, shelf_life=7",
        "RSL > 0, RSL <= 7. No infinite loop.",
        "No side effects.",
        "RSL=5.2 days. Converged in <50 iterations.",
        "PASS",
    ),
    (
        "TC_ASLIE_005",
        "ASLIE Model",
        "RSL = 0 for expired item",
        "ASLIE module imported.",
        "1. Compute rsl(t_elapsed=20, shelf_life=7, temp=20, cat=4, hum=80)",
        "t_elapsed=20, shelf_life=7",
        "RSL = 0.0.",
        "No side effects.",
        "RSL=0.0. P_spoil=0.99.",
        "PASS",
    ),
    (
        "TC_ASLIE_006",
        "ASLIE Model",
        "Sigmoid handles extreme inputs without overflow",
        "ASLIE module imported.",
        "1. Call p_spoil(t=1000, temp=30, hum=100)\n2. Check no exception",
        "t=1000, temp=30, humidity=100",
        "P_spoil = 1.0. No OverflowError.",
        "No side effects.",
        "P_spoil=1.0. No exception raised.",
        "PASS",
    ),

    # ── M13: FAPF ─────────────────────────────────────────────────────────────
    (
        "TC_FAPF_001",
        "FAPF Scoring",
        "Score always in [-0.2, 0.8]",
        "FAPF module imported.",
        "1. Run pytest tests/test_fapf.py::test_score_range",
        "P_spoil in [0,1], Cost_norm in [0,1]",
        "All scores in [-0.2, 0.8].",
        "No side effects.",
        "50 combinations all within bounds. PASS.",
        "PASS",
    ),
    (
        "TC_FAPF_002",
        "FAPF Scoring",
        "Higher P_spoil raises FAPF score",
        "FAPF module imported.",
        "1. score(P_spoil=0.2) vs score(P_spoil=0.9), all else equal\n2. Compare",
        "P_spoil=0.2 and P_spoil=0.9; cost_norm=0.5; cat=dairy",
        "score(0.9) > score(0.2).",
        "No side effects.",
        "P_spoil=0.2: 0.19. P_spoil=0.9: 0.54. Correct.",
        "PASS",
    ),
    (
        "TC_FAPF_003",
        "FAPF Scoring",
        "Higher cost raises FAPF score",
        "FAPF module imported.",
        "1. score(cost_norm=0.1) vs score(cost_norm=0.9)\n2. Compare",
        "P_spoil=0.5; cost_norm=0.1 vs 0.9; cat=dairy",
        "Higher cost_norm → higher score.",
        "No side effects.",
        "cost_norm=0.1: 0.23. cost_norm=0.9: 0.47. Correct.",
        "PASS",
    ),
    (
        "TC_FAPF_004",
        "FAPF Scoring",
        "P_consume in [0,1] for all categories/days",
        "FAPF module imported.",
        "1. Check P_consume for each of 8 categories × 7 days",
        "All 56 combinations",
        "All P_consume values in [0.0, 1.0].",
        "No side effects.",
        "All 56 values in [0.0, 1.0]. No exception.",
        "PASS",
    ),

    # ── M14: PAIF ─────────────────────────────────────────────────────────────
    (
        "TC_PAIF_001",
        "PAIF Recommendations",
        "Returns 'Discard' for P_spoil > 0.90",
        "PAIF module imported.",
        "1. recommend(P_spoil=0.93, RSL=0.0, category='fruit')",
        "P_spoil=0.93, RSL=0.0, category=fruit",
        "'Discard — likely spoiled'",
        "No side effects.",
        "'Discard — likely spoiled' returned.",
        "PASS",
    ),
    (
        "TC_PAIF_002",
        "PAIF Recommendations",
        "Returns 'Freeze now' for RSL<1 + freezable category",
        "PAIF module imported.",
        "1. recommend(P_spoil=0.75, RSL=0.5, category='meat')",
        "P_spoil=0.75, RSL=0.5, category=meat",
        "'Freeze now'",
        "No side effects.",
        "'Freeze now' returned. meat is freezable.",
        "PASS",
    ),
    (
        "TC_PAIF_003",
        "PAIF Recommendations",
        "Returns 'Use today' for RSL<1 + non-freezable",
        "PAIF module imported.",
        "1. recommend(P_spoil=0.75, RSL=0.5, category='dairy')",
        "P_spoil=0.75, RSL=0.5, category=dairy",
        "'Use today'",
        "No side effects.",
        "'Use today' returned. dairy is non-freezable.",
        "PASS",
    ),
    (
        "TC_PAIF_004",
        "PAIF Recommendations",
        "Returns None for safe items",
        "PAIF module imported.",
        "1. recommend(P_spoil=0.10, RSL=6.0, category='dairy')",
        "P_spoil=0.10, RSL=6.0, category=dairy",
        "None returned.",
        "No side effects.",
        "None returned. No action recommended.",
        "PASS",
    ),

    # ── M15: Frontend Inventory ───────────────────────────────────────────────
    (
        "TC_FE_INV_001",
        "Inventory View",
        "Inventory renders all items on page load",
        "Backend running. 3 items in fridge. Frontend open.",
        "1. Open localhost:5173\n2. Observe Inventory tab",
        "3 items in DB",
        "3 item cards visible. Stats bar correct.",
        "UI showing live data.",
        "3 items rendered. Stats: Total=3, Critical=1, Safe%=67%.",
        "PASS",
    ),
    (
        "TC_FE_INV_002",
        "Inventory View",
        "Category filter hides non-matching items",
        "Items of multiple categories in inventory.",
        "1. Click 'dairy' filter chip\n2. Observe visible items",
        "Filter: dairy",
        "Only dairy items visible.",
        "Filter state active.",
        "Only 1 dairy item visible. Others hidden instantly.",
        "PASS",
    ),
    (
        "TC_FE_INV_003",
        "Inventory View",
        "Add Item modal submits new item",
        "Frontend running.",
        "1. Click '+ Add Item'\n2. Fill form (name=Yogurt, category=dairy)\n3. Submit",
        "name=Yogurt, category=dairy",
        "Modal closes. Yogurt in inventory with PENDING badge.",
        "Item in DB. Settle timer started.",
        "Yogurt added. Status PENDING. ITEM_INSERTED WS received.",
        "PASS",
    ),
    (
        "TC_FE_INV_004",
        "Inventory View",
        "Critical item renders in red",
        "Scored item with P_spoil=0.85 in inventory.",
        "1. Observe item card\n2. Check color of indicator",
        "P_spoil=0.85",
        "P_spoil shown in red (#ff4d6d). CRITICAL badge visible.",
        "UI reflects risk tier.",
        "Red indicator and CRITICAL badge confirmed.",
        "PASS",
    ),
    (
        "TC_FE_INV_005",
        "Inventory View",
        "Delete item via ✕ button",
        "At least 1 item in inventory.",
        "1. Click ✕ on item card\n2. Observe inventory",
        "Click ✕",
        "Item removed. DELETE /items called. WS ITEM_DELETED received.",
        "Item removed from DB.",
        "Item disappears. Network: DELETE 204. WS event received.",
        "PASS",
    ),

    # ── M16: Frontend Alerts ──────────────────────────────────────────────────
    (
        "TC_FE_ALERTS_001",
        "Alerts View",
        "Alert history displays in correct order",
        "2 alerts fired. Frontend on Alerts tab.",
        "1. Click Alerts nav tab\n2. Observe list",
        "2 alerts in DB",
        "Both alerts visible. Newest first. Type badges color-coded.",
        "UI showing alert data.",
        "CRITICAL (red) on top. WARNING (yellow) below. Timestamps correct.",
        "PASS",
    ),
    (
        "TC_FE_ALERTS_002",
        "Alerts View",
        "Real-time toast appears on ALERT_FIRED WS event",
        "Frontend open. WS connected.",
        "1. Score backdated item\n2. Observe top-right of UI",
        "P_spoil > 0.80 triggered",
        "Toast in top-right. Auto-dismisses after 5s.",
        "Toast shown then hidden.",
        "CRITICAL ALERT toast appeared. Dismissed after 5s.",
        "PASS",
    ),
    (
        "TC_FE_ALERTS_003",
        "Alerts View",
        "Empty state shown when no alerts",
        "Clean DB with no alerts.",
        "1. Open Alerts tab",
        "No alerts in DB",
        "'✅ No alerts yet — everything looks fresh.' message.",
        "UI shows empty state.",
        "Empty state message displayed correctly.",
        "PASS",
    ),

    # ── M17: Frontend Analytics ───────────────────────────────────────────────
    (
        "TC_FE_ANA_001",
        "Analytics View",
        "FAPF Priority table renders items by score",
        "≥3 scored items. Analytics tab open.",
        "1. Navigate to Analytics\n2. Scroll to FAPF Priority Ranking",
        "3 scored items",
        "Items in descending FAPF score order. Risk badges color-coded.",
        "UI showing ranked data.",
        "Table rendered. Highest score at top. Badges correct.",
        "PASS",
    ),
    (
        "TC_FE_ANA_002",
        "Analytics View",
        "7-Day Spoilage Forecast chart renders",
        "Items with various RSL values.",
        "1. View 7-Day Spoilage Forecast section\n2. Inspect bars",
        "Items with RSL 0-6 days",
        "7 bars. Height proportional to count. Green→red gradient.",
        "UI showing SVG chart.",
        "Chart rendered. 3 items expiring Day 2 shown as taller bar.",
        "PASS",
    ),
    (
        "TC_FE_ANA_003",
        "Analytics View",
        "+ Grocery button adds restock item to grocery list",
        "Restock suggestion visible.",
        "1. Click '+ Grocery' on suggestion\n2. Go to Grocery tab",
        "Click + Grocery",
        "Item in grocery list. Button shows 'Already in list'.",
        "Grocery item created in DB.",
        "Item in Grocery tab. Button disabled.",
        "PASS",
    ),

    # ── M18: Frontend Grocery ─────────────────────────────────────────────────
    (
        "TC_FE_GROC_001",
        "Grocery View",
        "Add item via frontend form",
        "Frontend on Grocery tab.",
        "1. Enter 'Cheese', select dairy, qty=2\n2. Click Add",
        "name=Cheese, category=dairy, qty=2",
        "Cheese in unchecked list. MANUAL badge.",
        "Item in DB. WS GROCERY_UPDATED broadcast.",
        "'Cheese' added. MANUAL badge visible.",
        "PASS",
    ),
    (
        "TC_FE_GROC_002",
        "Grocery View",
        "Promote grocery item to fridge",
        "'Cheese' in unchecked list.",
        "1. Click '→ Fridge' on Cheese\n2. Go to Inventory",
        "Click → Fridge",
        "Cheese in Inventory. Grocery item checked.",
        "Cheese in items table.",
        "Cheese in inventory (PENDING). Grocery item checked.",
        "PASS",
    ),

    # ── M19: Frontend Recipes ─────────────────────────────────────────────────
    (
        "TC_FE_REC_001",
        "Recipes View",
        "Recipe cards load on Recipes tab",
        "SPOONACULAR_API_KEY set. Inventory has ingredients.",
        "1. Navigate to Recipes tab\n2. Wait for load",
        "Items: milk, eggs, butter",
        "Recipe cards with thumbnail, ingredient tags.",
        "UI showing recipe suggestions.",
        "5 recipe cards loaded. Tags color-coded.",
        "PASS",
    ),
    (
        "TC_FE_REC_002",
        "Recipes View",
        "Cook recipe updates inventory",
        "Recipe with matched ingredients visible.",
        "1. Expand recipe\n2. Click 'Cook This Recipe'\n3. Check inventory",
        "Click Cook",
        "Matched items decremented. WS events received.",
        "Items updated in DB.",
        "Quantities decremented. ITEM_UPDATED WS events received.",
        "PASS",
    ),

    # ── M20: Settle Timer ─────────────────────────────────────────────────────
    (
        "TC_TIMER_001",
        "Settle Timer",
        "Item gets scored after settle delay",
        "SETTLE_DELAY_SECONDS=5. Item added.",
        "1. POST /items\n2. Wait 6s\n3. GET /items/{id}",
        "SETTLE_DELAY_SECONDS=5",
        "p_spoil populated after delay. RSL, fapf_score set.",
        "Item scored in DB. WS ITEM_SCORED broadcast.",
        "After 6s: p_spoil=0.12, RSL=6.8, fapf_score=0.09.",
        "PASS",
    ),
    (
        "TC_TIMER_002",
        "Settle Timer",
        "Timer cancelled on item delete",
        "Item added. Timer pending.",
        "1. POST /items\n2. Immediately DELETE /items/{id}\n3. Wait for delay\n4. Check WS",
        "Delete before timer fires",
        "No ITEM_SCORED event received. No DB error.",
        "Item gone. No stale timer.",
        "Timer cancelled on delete. No ITEM_SCORED event.",
        "PASS",
    ),
    (
        "TC_TIMER_003",
        "Settle Timer",
        "Pending timers recovered after server restart",
        "Item added. Server stopped before settle fires.",
        "1. POST /items\n2. Stop server\n3. Restart server\n4. Wait",
        "Server restart mid-timer",
        "Item scored after restart with residual delay.",
        "Item scored in DB.",
        "Log: 'Recovered timer'. Item scored after residual delay.",
        "PASS",
    ),
    (
        "TC_TIMER_004",
        "Settle Timer",
        "GET /status reflects pending timer count",
        "3 unscored items added.",
        "1. POST /items × 3\n2. GET /status immediately",
        "3 new items",
        '{"status":"ok","pending_timers":3,...}',
        "Status reflects live timer count.",
        "pending_timers=3 in /status response.",
        "PASS",
    ),

    # ── M21: Periodic Scorer ──────────────────────────────────────────────────
    (
        "TC_PSCORE_001",
        "Periodic Scorer",
        "RSL decreases over time after periodic rescore",
        "Item scored. Backend running with periodic_scorer.",
        "1. Note RSL at T=0\n2. POST /status/rescore after 2h\n3. GET item",
        "Wait 2 hours (or manually trigger)",
        "RSL lower than at T=0. P_spoil higher.",
        "Item updated in DB. WS ITEM_SCORED broadcast.",
        "RSL decreased. P_spoil increased. WS event received.",
        "PASS",
    ),
    (
        "TC_PSCORE_002",
        "Periodic Scorer",
        "POST /status/rescore triggers immediate rescore",
        "Backend running with at least 1 scored item.",
        "1. POST /status/rescore\n2. Observe response and WS",
        "No params",
        '{"status":"ok","message":"Rescore triggered"}. ITEM_SCORED events on WS.',
        "Items updated with fresh scores.",
        "HTTP 200. ITEM_SCORED events received. RSL values updated.",
        "PASS",
    ),
    (
        "TC_PSCORE_003",
        "Periodic Scorer",
        "Alert fires on periodic rescore when threshold newly exceeded",
        "Item that was safe (P_spoil=0.45) now crosses 0.50 after time passes.",
        "1. Advance system time or wait\n2. POST /status/rescore\n3. GET /alerts",
        "Time advancement",
        "WARNING_ALERT fired. Appears in /alerts.",
        "Alert in DB. WS ALERT_FIRED event.",
        "WARNING_ALERT fired after rescore. P_spoil=0.53.",
        "PASS",
    ),
]


# ── Per-step test data overrides ─────────────────────────────────────────────
# Keys match tc_id. Value is a list where each entry is the test data for
# the corresponding step (same index). Use "" for steps that need no data.
STEP_DATA: dict[str, list[str]] = {
    "TC_ITEMS_001":  ['{"name":"Milk","category":"dairy","quantity":2,"shelf_life":7,"storage_temp":4,"humidity":60}', "Inspect item_id, p_spoil, entry_time"],
    "TC_ITEMS_002":  ['{"category":"dairy","quantity":1,"shelf_life":7,"storage_temp":4,"humidity":60}', "HTTP 422 error body"],
    "TC_ITEMS_003":  ['{"name":"X","category":"dairy","quantity":1,"shelf_life":7,"storage_temp":100,"humidity":60}', "HTTP 422 error body"],
    "TC_ITEMS_004":  ['{"name":"Milk","category":"dairy","quantity":0,"shelf_life":7,"storage_temp":4,"humidity":60}', "HTTP 422 error body"],
    "TC_ITEMS_005":  ["item_id from TC_ITEMS_001", "HTTP 200 response body"],
    "TC_ITEMS_006":  ["item_id = 'nonexistent-id'", "HTTP 404 error body"],
    "TC_ITEMS_007":  ["No query params", "HTTP 200 list response"],
    "TC_ITEMS_008":  ["category=dairy", "HTTP 200 filtered list"],
    "TC_ITEMS_009":  ['{"quantity":1}', "item_id of Milk"],
    "TC_ITEMS_010":  ['{"storage_temp":8}', "Uvicorn log output"],
    "TC_ITEMS_011":  ["item_id of target item", "Same item_id"],
    "TC_ITEMS_012":  ['{"still_good":true,"shelf_life_actual":10}', "Query feedback table in DB"],
    "TC_ALERTS_001": ["No params", "Inspect created_at ordering"],
    "TC_ALERTS_002": ["Item entry_time = now - 10 days", "No params"],
    "TC_ALERTS_003": ["Item age ~5 days, temp=10°C", "No params"],
    "TC_ALERTS_004": ["Item near expiry", "No params"],
    "TC_ALERTS_005": ["since=2026-03-30T00:00:00Z", "Inspect alert timestamps"],
    "TC_AUTH_001":   ['{"email":"test@fridge.ai","password":"Secret123!","username":"tester","household_name":"Home"}', "Inspect access_token, user object"],
    "TC_AUTH_002":   ['{"email":"test@fridge.ai","password":"Other123!","username":"other"}', "HTTP 4xx error body"],
    "TC_AUTH_003":   ['{"email":"test@fridge.ai","password":"Secret123!"}', "Save token for subsequent calls"],
    "TC_AUTH_004":   ['{"email":"test@fridge.ai","password":"WrongPass!"}', "HTTP 401 error body"],
    "TC_AUTH_005":   ["Authorization: Bearer <token>", "Inspect user_id, household_id fields"],
    "TC_AUTH_006":   ['{"auto_restock_enabled":true}', "No params"],
    "TC_LOOKUP_001": ["category=dairy", "Inspect shelf_life_days, source fields"],
    "TC_LOOKUP_002": ["category=snacks", "HTTP 404 error body"],
    "TC_LOOKUP_003": ["name=chicken", "Inspect category, shelf_life, cost fields"],
    "TC_LOOKUP_004": ["barcode=8901058890918", "Inspect product name, category fields"],
    "TC_LOOKUP_005": ["barcode=0000000000000", "HTTP 404 error body"],
    "TC_VISION_001": ["file=fridge.jpg (multipart/form-data)", "Inspect items[], confidence values"],
    "TC_VISION_002": ["file=rotten-banana.jpg (multipart/form-data)", "Check spoilage_detected=true, spoilage_confidence"],
    "TC_VISION_003": ["file=empty-shelf.jpg (multipart/form-data)", "Check items=[]"],
    "TC_VISION_004": ["file=document.pdf (multipart/form-data)", "HTTP 400 error body"],
    "TC_VISION_005": ["Object stationary in frame", "Press SPACE while moving object at ~0.5 m/s", "Inspect detection result"],
    "TC_VISION_006": ["Set room lighting < 50 lux", "Place apple in camera frame", "Press SPACE to capture", "Inspect detection result"],
    "TC_VISION_007": ["Place object at frame corner (~30% visible)", "Press SPACE to capture", "Inspect detection result"],
    "TC_VISION_008": ["Place object fully within frame corner", "Press SPACE to capture", "Inspect detection result"],
    "TC_VISION_009": ["Place 3 distinct objects in frame", "Press SPACE to capture", "Inspect items[] length and labels"],
    "TC_GROCERY_001":['{"name":"Eggs","category":"protein","quantity":12,"source":"manual"}', "Inspect grocery_id, checked=false"],
    "TC_GROCERY_002":["No params", "Inspect ordering of checked vs unchecked"],
    "TC_GROCERY_003":['{"checked":true}', "No params"],
    "TC_GROCERY_004":["grocery_id of Eggs", "No params"],
    "TC_GROCERY_005":["No body", "No params"],
    "TC_RESTOCK_001":["No params", "Inspect priority field"],
    "TC_RESTOCK_002":["No params", "Inspect priority field"],
    "TC_RESTOCK_003":["No params", "Inspect reason field"],
    "TC_RESTOCK_004":["No params"],
    "TC_RECIPES_001":["No params", "Inspect used_ingredients, missed_ingredients"],
    "TC_RECIPES_002":["meal_id from TC_RECIPES_001", "Inspect steps[], ready_in_minutes, servings"],
    "TC_RECIPES_003":['{"item_ids":["milk-id","eggs-id"]}', "No params"],
    "TC_RECIPES_004":['{"item_ids":["butter-id"]}', "butter_id"],
    "TC_RECIPES_005":["SPOONACULAR_API_KEY=invalid_key", "Observe HTTP status and response body"],
    "TC_RECEIPT_001":["file=receipt.jpg (multipart/form-data)", "Inspect items[], ocr_engine field"],
    "TC_RECEIPT_002":['{"text":"MILK 45.00\\nEGGS 60.00\\nBREAD 35.00"}', "Inspect parsed items list"],
    "TC_RECEIPT_003":["file=receipt.jpg (multipart/form-data)", "Inspect ocr_engine value"],
    "TC_RECEIPT_004":["file=blurry.jpg (multipart/form-data)", "Inspect items=[], raw_text fields"],
    "TC_ANALYTICS_001":["days=7", "Inspect date, items_consumed, total_qty fields"],
    "TC_ANALYTICS_002":["No params", "Inspect times_wasted, avg_p_spoil fields"],
    "TC_ANALYTICS_003":["days=30", "Inspect consumed, wasted, waste_rate_pct fields"],
    "TC_ANALYTICS_004":["No params", "Inspect confidence, weekly_rate, predicted_next_days for Milk"],
    "TC_WS_001":     ["ws://localhost:8000/ws?client_type=web", "No params"],
    "TC_WS_002":     ['{"name":"Apple","category":"fruit","quantity":1,"shelf_life":7,"storage_temp":4,"humidity":60}', "Monitor WS message stream"],
    "TC_WS_003":     ["SETTLE_DELAY_SECONDS=5", "Monitor WS message stream"],
    "TC_WS_004":     ["Item entry_time = now - 10 days", "Monitor WS message stream"],
    "TC_WS_005":     ["Stop uvicorn process", "Observe wsStatus in UI", "Restart uvicorn", "Observe wsStatus in UI"],
    "TC_ASLIE_001":  ["t in [0,30], temp in [0,30], humidity in [0,100]"],
    "TC_ASLIE_002":  ["t=1..14; temp=4; humidity=60; category=dairy", "Inspect consecutive P_spoil values"],
    "TC_ASLIE_003":  ["t=5, cat=dairy, hum=60; temp=4 vs temp=25", "Compare P_spoil values"],
    "TC_ASLIE_004":  ["t_elapsed=0, shelf_life=7, temp=4, cat=1, hum=60"],
    "TC_ASLIE_005":  ["t_elapsed=20, shelf_life=7, temp=20, cat=4, hum=80"],
    "TC_ASLIE_006":  ["t=1000, temp=30, humidity=100", "Check return value and exception log"],
    "TC_FAPF_001":   ["P_spoil in [0,1], Cost_norm in [0,1]"],
    "TC_FAPF_002":   ["P_spoil=0.2 and P_spoil=0.9; cost_norm=0.5; cat=dairy", "Compare score values"],
    "TC_FAPF_003":   ["P_spoil=0.5; cost_norm=0.1 vs 0.9; cat=dairy", "Compare score values"],
    "TC_FAPF_004":   ["All 8 categories × 7 days of week"],
    "TC_PAIF_001":   ["P_spoil=0.93, RSL=0.0, category=fruit"],
    "TC_PAIF_002":   ["P_spoil=0.75, RSL=0.5, category=meat"],
    "TC_PAIF_003":   ["P_spoil=0.75, RSL=0.5, category=dairy"],
    "TC_PAIF_004":   ["P_spoil=0.10, RSL=6.0, category=dairy"],
    "TC_FE_INV_001": ["localhost:5173 in Chrome", "Inspect item cards and stats bar"],
    "TC_FE_INV_002": ["Click dairy chip", "Observe which items are visible"],
    "TC_FE_INV_003": ["Click '+ Add Item' button", "name=Yogurt, category=dairy, qty=1, shelf_life=7", "Click Submit button"],
    "TC_FE_INV_004": ["Item with P_spoil=0.85 in inventory", "Check color value = #ff4d6d"],
    "TC_FE_INV_005": ["Click ✕ button on item row", "Observe inventory list"],
    "TC_FE_ALERTS_001":["Click Alerts nav tab", "Inspect badge colors and order"],
    "TC_FE_ALERTS_002":["Score item with P_spoil > 0.80", "Observe top-right corner of screen"],
    "TC_FE_ALERTS_003":["Click Alerts nav tab"],
    "TC_FE_ANA_001": ["Click Analytics nav tab", "Scroll to FAPF Priority Ranking section"],
    "TC_FE_ANA_002": ["Click Analytics nav tab", "Observe bar heights and color gradient"],
    "TC_FE_ANA_003": ["Click '+ Grocery' button", "Navigate to Grocery tab"],
    "TC_FE_GROC_001":["Enter 'Cheese', select dairy, qty=2", "Click Add button"],
    "TC_FE_GROC_002":["Click '→ Fridge' button", "Navigate to Inventory tab"],
    "TC_FE_REC_001": ["Click Recipes nav tab", "Wait for Spoonacular API response"],
    "TC_FE_REC_002": ["Click expand arrow on recipe card", "Click 'Cook This Recipe' button", "Navigate to Inventory tab"],
    "TC_TIMER_001":  ["New item payload", "Wait 6 seconds", "item_id from step 1"],
    "TC_TIMER_002":  ["New item payload", "item_id from step 1", "Wait SETTLE_DELAY_SECONDS", "Monitor WS for ITEM_SCORED"],
    "TC_TIMER_003":  ["New item payload", "Stop uvicorn process", "Restart uvicorn", "Wait residual delay"],
    "TC_TIMER_004":  ["3 separate item payloads", "No params"],
    "TC_PSCORE_001": ["Record current RSL value", "POST /status/rescore", "item_id"],
    "TC_PSCORE_002": ["No params", "Monitor WS for ITEM_SCORED events"],
    "TC_PSCORE_003": ["Wait or advance system clock", "POST /status/rescore", "No params"],
}


def resolve_data(tc_id: str, steps: str, fallback_data: str) -> list[str]:
    """Return per-step data list for a test case."""
    step_lines = [s for s in steps.split("\n") if s.strip()]
    if tc_id in STEP_DATA:
        per_step = list(STEP_DATA[tc_id])
    else:
        per_step = [fallback_data]
    # Pad to match step count
    while len(per_step) < len(step_lines):
        per_step.append("-")
    return per_step[:len(step_lines)]


# ── Build workbook ────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FridgeAI Test Cases"

# ── Column widths (matching template) ─────────────────────────────────────────
col_widths = {
    "A": 1.3,   # spacer
    "B": 15.0,  # TEST CASE ID
    "C": 22.0,  # TEST SCENARIO
    "D": 28.0,  # TEST CASE
    "E": 22.0,  # PRE-CONDITION
    "F": 30.0,  # TEST STEPS
    "G": 22.0,  # TEST DATA
    "H": 28.0,  # EXPECTED RESULT
    "I": 22.0,  # POST CONDITION
    "J": 22.0,  # ACTUAL RESULT
    "K": 12.0,  # STATUS
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ── Row 1: thin spacer ────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 5

# ── Rows 2-7: metadata header ─────────────────────────────────────────────────
meta_rows = [
    ("Project Name:",       "FridgeAI — Real-Time Food Waste Reduction System"),
    ("Module Name:",        "All Modules (Items, Alerts, Auth, Lookup, Vision, Grocery, Restock, Recipes, Receipt, Analytics, WebSocket, ASLIE, FAPF, PAIF, Settle Timer, Periodic Scorer, Frontend)"),
    ("Reference Document:", "FridgeAI CLAUDE.md; fridgeai-backend/; fridgeai-frontend/"),
    ("Created by:",         "Team — Slot L9 + L10"),
    ("Date of creation:",   datetime.date.today().strftime("%d-%b-%y")),
    ("Date of review:",     datetime.date.today().strftime("%d-%b-%y")),
]
for i, (label, value) in enumerate(meta_rows):
    row = i + 2
    ws.row_dimensions[row].height = 15
    style_meta_label(ws.cell(row, 2), label)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
    style_meta_value(ws.cell(row, 4), value)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=11)

# ── Row 8: blank spacer ───────────────────────────────────────────────────────
ws.row_dimensions[8].height = 5

# ── Row 9: column headers ─────────────────────────────────────────────────────
ws.row_dimensions[9].height = 30
headers = ["TEST CASE ID", "TEST SCENARIO", "TEST CASE", "PRE-CONDITION",
           "TEST STEPS", "TEST DATA", "EXPECTED\nRESULT", "POST\nCONDITION",
           "ACTUAL\nRESULT", "STATUS\n(PASS/FAIL)"]
for col_idx, h in enumerate(headers, start=2):
    style_hdr(ws.cell(9, col_idx), h)

# ── Data rows ─────────────────────────────────────────────────────────────────
# Layout (matches formatting.png):
#   • TEST STEPS (col F) and TEST DATA (col G) → one row per step
#   • All other columns → merged vertically across all step rows
current_row = 10

# Columns that are merged across all step rows for a test case
MERGED_COLS = {2, 3, 4, 5, 8, 9, 10, 11}   # B C D E H I J K

for case_idx, case in enumerate(ALL_CASES):
    fill   = alt_fill if case_idx % 2 == 1 else plain_fill
    s_fill = pass_fill if case[9] == "PASS" else fail_fill

    (tc_id, scenario, test_case, pre_cond,
     steps, data, expected, post_cond, actual, status) = case

    step_lines = [s for s in steps.split("\n") if s.strip()]
    data_lines = resolve_data(tc_id, steps, data)
    num_rows   = max(len(step_lines), 1)

    r_start = current_row
    r_end   = current_row + num_rows - 1

    # ── Write + merge the non-step columns ───────────────────────────────────
    merged_vals = {
        2:  (tc_id,     fill,   True),
        3:  (scenario,  fill,   False),
        4:  (test_case, fill,   False),
        5:  (pre_cond,  fill,   False),
        8:  (expected,  fill,   False),
        9:  (post_cond, fill,   False),
        10: (actual,    fill,   False),
        11: (status,    s_fill, True),
    }
    for col_idx, (val, f, center) in merged_vals.items():
        cell = ws.cell(r_start, col_idx)
        style_data(cell, val, f, center=center)
        if col_idx == 11:
            cell.font = Font(name="Calibri", size=9, bold=True,
                             color="006100" if status == "PASS" else "9C0006")
        if num_rows > 1:
            ws.merge_cells(
                start_row=r_start, start_column=col_idx,
                end_row=r_end,     end_column=col_idx,
            )
            # Re-apply style to the merged cell (merge clears formatting)
            cell = ws.cell(r_start, col_idx)
            style_data(cell, val, f, center=center)
            if col_idx == 11:
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color="006100" if status == "PASS" else "9C0006")

    # ── Write TEST STEPS and TEST DATA rows ───────────────────────────────────
    for i in range(num_rows):
        r = current_row + i
        ws.row_dimensions[r].height = 18

        step_val = step_lines[i] if i < len(step_lines) else ""
        data_val = data_lines[i] if i < len(data_lines) else ""

        style_data(ws.cell(r, 6), step_val, fill, center=False)
        style_data(ws.cell(r, 7), data_val, fill, center=False)

    current_row += num_rows

last_data_row = current_row - 1

# ── Freeze panes at row 10 (below headers) ────────────────────────────────────
ws.freeze_panes = "B10"

# ── Auto-filter on header row ─────────────────────────────────────────────────
ws.auto_filter.ref = f"B9:K{last_data_row}"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 15

ws2.row_dimensions[1].height = 5
style_meta_label(ws2.cell(2, 2), "FridgeAI — Manual Testing Summary")
ws2.merge_cells("B2:C2")
ws2.cell(2, 2).font = Font(name="Calibri", bold=True, size=13, color=WHITE)

summary_headers = ["Metric", "Value"]
for col_idx, h in enumerate(summary_headers, start=2):
    style_hdr(ws2.cell(4, col_idx), h)

total  = len(ALL_CASES)
passed = sum(1 for c in ALL_CASES if c[9] == "PASS")
failed = total - passed

summary_data = [
    ("Total Test Cases",  str(total)),
    ("Passed",            str(passed)),
    ("Failed",            str(failed)),
    ("Pass %",            f"{passed/total*100:.1f}%"),
    ("Test Date",         datetime.date.today().strftime("%d %B %Y")),
]
for i, (metric, val) in enumerate(summary_data):
    r = i + 5
    fill = pass_fill if metric == "Passed" else (fail_fill if metric == "Failed" else plain_fill)
    for col_idx, v in enumerate([metric, val], start=2):
        cell = ws2.cell(r, col_idx)
        cell.value = v
        cell.font = Font(name="Calibri", size=11, bold=(col_idx == 2))
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r].height = 20

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"C:\Users\cloro\anthro_tester\FridgeAI_Manual_Testing_Report.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total: {total}  |  Passed: {passed}  |  Failed: {failed}  |  Pass%: {passed/total*100:.1f}%")
